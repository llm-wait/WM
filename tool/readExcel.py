import os
import re

import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, colors
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import ColorList

from WM.tool.project_path import test_data_path
from WM.tool.readconfig import ReadConfig


class ReadExcel():
    def __init__(self, filename):
        self.filename = filename
        self.wb = openpyxl.load_workbook(test_data_path + "/" + filename)
        self.confile="wmcase.conf"

    def read_cases(self):
        '''
            从配置文件case.conf中读取[MODE]->mode值,(测试的工作表和执行的模式)
            mode控制是否执行所有的用例，mode值只能是all或列表，默认值all表示执行所有的用例，
            列表表示执行指定的用例
        '''
        mode= eval(ReadConfig(self.confile).read_config("MODE", 'mode'))
        test_data=[]
        for key in mode.keys():
            sheet = self.wb[key]
            if mode[key]=="all":  #all值为执行所有用例
                for i in range(2, sheet.max_row + 1):
                    # tel=getattr(GetData, "LoginId")
                    tel = '18031443065'
                    # print("******",tel)
                    row_data = {}  # 单行测试用例
                    row_data["case_id"] = sheet.cell(i, 1).value
                    row_data["module"] = sheet.cell(i, 2).value
                    row_data["url"] = sheet.cell(i, 3).value
                    row_data["method"] = sheet.cell(i, 4).value
                    row_data["title"] = sheet.cell(i, 5).value
                    row_data["params"] = sheet.cell(i, 6).value
                    # if  sheet.cell(i, 6).value.find('${tel_num}')!=-1:
                    #     row_data["params"] = sheet.cell(i, 6).value.replace('${tel_num}',tel)
                    # else:
                    #     row_data["params"]=sheet.cell(i, 6).value
                    row_data["header"] = sheet.cell(i, 7).value
                    # 依赖的接口case_id
                    row_data["bedepnedid"] = sheet.cell(i, 8).value
                    # 被依赖的接口返回值
                    row_data["bedepnedkey"] = sheet.cell(i, 9).value
                    # 被依赖的字段
                    row_data["depnedkey"] = sheet.cell(i, 10).value
                    row_data["depneddata"] = sheet.cell(i, 11).value
                    row_data["expected"] = sheet.cell(i, 12).value
                    row_data["responsedata"] = sheet.cell(i, 13).value
                    # 测试用例中添加sheetname字典，用于测试结果写回时之指定excel
                    row_data["sheetname"] = key
                    test_data.append(row_data)

            else:
                for i  in eval(mode[key]):
                    row_data = {}  # 单行测试用例
                    row_data["case_id"] = sheet.cell(i+1, 1).value
                    row_data["module"] = sheet.cell(i+1, 2).value
                    row_data["url"] = sheet.cell(i+1, 3).value
                    row_data["method"] = sheet.cell(i+1, 4).value
                    row_data["title"] = sheet.cell(i+1, 5).value
                    row_data["params"] = sheet.cell(i+1, 6).value
                    row_data["header"] = sheet.cell(i+1, 7).value
                    # 依赖的接口case_id
                    row_data["bedepnedid"] = sheet.cell(i+1, 8).value
                    # 被依赖的接口返回值
                    row_data["bedepnedkey"] = sheet.cell(i+1, 9).value
                    # 被依赖的字段
                    row_data["depnedkey"] = sheet.cell(i+1, 10).value
                    row_data["depneddata"] = sheet.cell(i+1, 13).value
                    row_data["expected"] = sheet.cell(i+1, 12).value
                    row_data["responsedata"] = sheet.cell(i, 13).value
                    # 测试用例中添加sheetname字典，用于测试结果写回时之指定excel
                    row_data["sheetname"] = key
                    test_data.append(row_data)
        return test_data

    def write_back(self, sheet_name, i, j, value, back_color=None):
            '''
            用于写回测试结果
            '''
            sheet = self.wb[sheet_name]
            sheet.cell(i, j).value = value
            # 设置写回单元格填充颜色,默认是白色
            if back_color == None:
                sheet.cell(i, j).fill = PatternFill("solid", fgColor="FFFFFF")
            else:
                sheet.cell(i, j).fill = PatternFill("solid", fgColor=back_color)
            self.wb.save(test_data_path + "/" + self.filename)

    def write_back_dependdata(self, sheet_name, row, data):
            sheet = self.wb[sheet_name]
            sheet.cell(row, 11).value = data
            self.wb.save(test_data_path + "/" + self.filename)

    def write_back_params(self, sheet_name, row, data):
            sheet = self.wb[sheet_name]
            sheet.cell(row, 6).value = data
            self.wb.save(test_data_path + "/" + self.filename)

    def read_dependcase(self, sheetname, case_id):
            '''
            用于读取接口依赖测试数据
            根据传递的参数case_id,找到对应行的数据
            :param sheetname:被依赖的接口所在工作表
            :param  case_id:被依赖的接口，是个int
            :return:
            '''
            sheet = self.wb[sheetname]
            depandcase = []
            for i in range(1, sheet.max_column + 1):
                depandcase.append(sheet.cell(int(case_id.split("_")[1]) + 1, i).value)
            return depandcase

    def read_dependcase1(self, sheetname,rownum):
            '''
            用于读取接口依赖测试数据
            根据传递的参数case_id,找到对应行的数据
            :param sheetname:被依赖的接口所在工作表
            :return:
            '''
            sheet = self.wb[sheetname]

            data=sheet.cell(rownum,1).value
            return data

    def read_handleparam(self, sheetname, rownum):
        '''
        用于读取处理后的请求参数
        :param sheetname: 工作表名称
        :param rownum: 数据所在行数
        :return:
        '''
        sheet = self.wb[sheetname]

        data = sheet.cell(rownum, 6).value
        return data


if __name__ == '__main__':
    readex = ReadExcel("批量业务接口.xlsx")
    before_data = readex.read_cases()
    print(before_data)
    # dependkey=readex.read_dependcase( "登录（已调整）","case_004")
    # print(dependkey)
    # print(readex.read_dependcase1("init", 2))